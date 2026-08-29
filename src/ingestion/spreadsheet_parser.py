"""CSV / XLSX ingestion for investor target lists and CRM exports.

The header row is rarely where a naive reader expects it and column names differ between
every CRM. This module finds the header row, normalises the headings to canonical keys,
and keeps the original values untouched alongside them.
"""

from __future__ import annotations

import csv
import io
from pathlib import Path

from ..models.evidence import SourceType, Warning
from ..utils.logging import get_logger
from ..utils.text import squeeze
from .types import ParsedDocument, Segment, TableRow, UnreadableFile

_log = get_logger()

#: Canonical key -> heading fragments seen in real target lists and CRM exports.
COLUMN_SYNONYMS: dict[str, tuple[str, ...]] = {
    "investor_name": (
        "investor",
        "fund",
        "firm",
        "organization",
        "organisation",
        "company",
        "account",
        "name",
        "prospect",
        "vc",
    ),
    "investor_type": ("type", "category", "investor type", "entity type", "class"),
    "check_size": (
        "check",
        "cheque",
        "check size",
        "typical check",
        "allocation",
        "target check",
        "investment size",
        "ticket",
    ),
    "stage_focus": ("stage", "stage focus", "stages", "entry stage", "round stage"),
    "sector_focus": ("sector", "industry", "vertical", "focus", "thesis", "sector focus"),
    "leads_rounds": ("leads", "leads rounds", "can lead", "lead?", "lead capable", "lead"),
    "lead_history": ("lead history", "led rounds", "led", "lead evidence", "led deals"),
    "relationship": (
        "relationship",
        "warmth",
        "connection",
        "intro",
        "intro path",
        "warm intro",
        "source of intro",
        "referral",
    ),
    "status": (
        "status",
        "stage in process",
        "pipeline stage",
        "deal stage",
        "diligence",
        "current status",
        "process stage",
    ),
    "contact": ("contact", "partner", "person", "poc", "champion", "owner contact"),
    "notes": ("notes", "comment", "comments", "detail", "details", "summary"),
    "portfolio": ("portfolio", "portfolio companies", "investments", "relevant portfolio"),
    "fund_status": ("fund status", "deployment", "fund", "vintage", "fund vintage"),
    "committed": ("committed", "commitment", "soft circle", "circled", "amount"),
    "next_step": ("next step", "action", "next action", "todo", "follow up", "follow-up"),
    "owner": ("owner", "responsible", "assigned", "lead owner"),
    "location": ("location", "geography", "region", "hq", "city", "country"),
    "aum": ("aum", "fund size", "assets"),
    "website": ("website", "url", "domain", "link"),
}


def parse_spreadsheet(path: str | Path, source_type: SourceType = SourceType.INVESTOR_LIST) -> ParsedDocument:
    path = Path(path)
    suffix = path.suffix.lower()
    if suffix == ".csv":
        sheets = {"csv": _read_csv(path)}
    elif suffix in {".xlsx", ".xlsm"}:
        sheets = _read_xlsx(path)
    else:
        raise UnreadableFile(f"{path.name} is not a spreadsheet this reader supports.")

    doc = ParsedDocument(path=path, name=path.name, source_type=source_type, kind="spreadsheet")

    index = 0
    for sheet_name, grid in sheets.items():
        index += 1
        grid = [row for row in grid if any(squeeze(str(c)) for c in row)]
        if not grid:
            doc.warnings.append(
                Warning(
                    severity="info",
                    stage="ingestion",
                    message=f"Sheet '{sheet_name}' in {path.name} is empty.",
                )
            )
            continue

        header_index = _find_header_row(grid)
        headers_raw = [squeeze(str(c)) for c in grid[header_index]]
        mapping = normalise_headers(headers_raw)
        doc.headers = [h for h in headers_raw if h]

        text_lines = [" | ".join(h for h in headers_raw if h)]
        for offset, raw_row in enumerate(grid[header_index + 1 :], start=header_index + 2):
            values: dict[str, str] = {}
            for column, cell in enumerate(raw_row):
                value = squeeze(str(cell)) if cell is not None else ""
                if not value:
                    continue
                heading = headers_raw[column] if column < len(headers_raw) else f"column_{column}"
                values[heading] = value
                canonical = mapping.get(column)
                if canonical and canonical not in values:
                    values[canonical] = value
            if not values:
                continue
            doc.rows.append(TableRow(sheet=sheet_name, row_number=offset, values=values))
            text_lines.append(" | ".join(f"{k}: {v}" for k, v in values.items() if k in headers_raw))

        doc.segments.append(Segment(index=index, text="\n".join(text_lines), kind="sheet", title=sheet_name))

    if not doc.rows:
        doc.warnings.append(
            Warning(
                severity="warning",
                stage="ingestion",
                message=f"No data rows were found in {path.name}.",
                detail="Check that the file has a header row followed by one row per investor.",
            )
        )
    if doc.rows and not any("investor_name" in row.values for row in doc.rows):
        doc.warnings.append(
            Warning(
                severity="warning",
                stage="ingestion",
                message=(
                    f"No investor-name column was recognised in {path.name}; the first text "
                    "column will be used."
                ),
                detail=f"Headings seen: {', '.join(doc.headers[:12])}",
            )
        )
    doc.metadata = {"rows": len(doc.rows), "sheets": len(sheets)}
    return doc


def normalise_headers(headers: list[str]) -> dict[int, str]:
    """Map column index -> canonical key. Unrecognised columns are left out."""
    mapping: dict[int, str] = {}
    used: set[str] = set()
    lowered = [h.strip().lower() for h in headers]

    # Exact-ish matches first so that "Fund Status" does not claim the "fund" slot before
    # "Investor Name" has had a chance at it.
    for canonical, synonyms in COLUMN_SYNONYMS.items():
        for column, heading in enumerate(lowered):
            if column in mapping or canonical in used or not heading:
                continue
            if heading in synonyms:
                mapping[column] = canonical
                used.add(canonical)
    for canonical, synonyms in COLUMN_SYNONYMS.items():
        if canonical in used:
            continue
        for column, heading in enumerate(lowered):
            if column in mapping or not heading:
                continue
            if any(s in heading for s in synonyms):
                mapping[column] = canonical
                used.add(canonical)
                break
    return mapping


def _find_header_row(grid: list[list]) -> int:
    """The header is the first row within the top few whose cells are mostly short text."""
    best_index, best_score = 0, -1.0
    for index, row in enumerate(grid[:6]):
        cells = [squeeze(str(c)) for c in row if squeeze(str(c))]
        if len(cells) < 2:
            continue
        texty = sum(1 for c in cells if not c.replace(",", "").replace(".", "").isdigit())
        short = sum(1 for c in cells if len(c) <= 40)
        recognised = sum(
            1 for c in cells if any(s in c.lower() for syns in COLUMN_SYNONYMS.values() for s in syns)
        )
        score = texty + short * 0.5 + recognised * 2 - index * 0.5
        if score > best_score:
            best_index, best_score = index, score
    return best_index


def _read_csv(path: Path) -> list[list[str]]:
    raw = path.read_bytes()
    for encoding in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
        try:
            text = raw.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    else:  # pragma: no cover - latin-1 always decodes
        raise UnreadableFile(f"{path.name} could not be decoded as text.")

    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
    except csv.Error:
        dialect = csv.excel
    return [row for row in csv.reader(io.StringIO(text), dialect)]


def _read_xlsx(path: Path) -> dict[str, list[list]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - dependency is declared
        raise UnreadableFile("openpyxl is not installed.") from exc
    try:
        workbook = load_workbook(str(path), data_only=True, read_only=True)
    except Exception as exc:
        raise UnreadableFile(f"{path.name} could not be opened as a workbook: {exc}") from exc

    sheets: dict[str, list[list]] = {}
    for worksheet in workbook.worksheets:
        rows = []
        for row in worksheet.iter_rows(values_only=True):
            rows.append(["" if cell is None else cell for cell in row])
        sheets[worksheet.title] = rows
    workbook.close()
    return sheets
